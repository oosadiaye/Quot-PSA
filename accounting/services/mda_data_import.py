"""
Bulk data import for MDA submissions at the OAGF.

The Office of the Accountant-General consumes monthly data from
MDAs — journal summaries, revenue collection schedules, provision
registers — typically delivered as spreadsheets. This service parses
CSV / XLSX uploads into structured rows the AG then validates and
consolidates.

Design goals
------------
1. **Tolerant parsing**: CSVs come from Excel (BOM, CRLF), Google
   Sheets (UTF-8, LF), legacy systems (cp1252). All three decode
   paths are tried. XLSX reads via openpyxl.
2. **Header-driven**: the first non-empty row defines the columns.
   Downstream consumers match on header names, not positions, so a
   spreadsheet with an extra column at the start still works.
3. **No side effects**: the importer returns structured rows + errors
   but does NOT persist to any model. The caller (view / task) decides
   what to do with the parsed data — validate, preview, consolidate,
   dry-run.
4. **Schema-light**: callers supply a ``spec`` describing required and
   optional columns + coercion hints. Wrong column name → row-level
   error, not a 500.

Usage
-----
    from accounting.services.mda_data_import import MDAImporter, ImportSpec

    spec = ImportSpec(
        required_columns=['date', 'mda_code', 'amount', 'narration'],
        numeric_columns=['amount'],
        date_columns=['date'],
    )
    result = MDAImporter.parse(uploaded_file, spec)
    if result.errors:
        return Response({'errors': result.errors}, status=400)
    return Response({'rows': result.rows})
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any


# ── Public dataclasses ─────────────────────────────────────────────────

@dataclass(frozen=True)
class ImportSpec:
    """Describes what columns the importer expects.

    ``required_columns``: row must have these, non-empty.
    ``optional_columns``: informational only; unknown columns in the
        file are preserved as raw strings.
    ``numeric_columns``:  values parsed via ``_parse_decimal``.
    ``date_columns``:     values parsed via ``_parse_date``.
    ``max_rows``:         defensive cap so a runaway upload can't
        exhaust memory. Default 50k rows.
    """
    required_columns: list[str] = field(default_factory=list)
    optional_columns: list[str] = field(default_factory=list)
    numeric_columns: list[str] = field(default_factory=list)
    date_columns: list[str] = field(default_factory=list)
    max_rows: int = 50_000


@dataclass
class ImportResult:
    """Parsed + validated output."""
    rows: list[dict[str, Any]]
    errors: list[dict[str, Any]]
    # Header as seen in the file (post-normalisation).
    columns: list[str]
    # Raw stats for UI display.
    total_rows: int
    accepted_rows: int
    rejected_rows: int

    def is_valid(self) -> bool:
        return not self.errors and self.accepted_rows > 0


# ── Main service ───────────────────────────────────────────────────────

class MDAImporter:
    """Parses CSV / XLSX uploads into structured rows."""

    CSV_ENCODINGS = ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1')
    DATE_FORMATS = (
        '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
        '%d %b %Y', '%d-%b-%Y',
    )

    @classmethod
    def parse(cls, uploaded_file, spec: ImportSpec) -> ImportResult:
        """Dispatch to the right parser based on filename extension.

        Treats anything without a ``.xlsx`` / ``.xlsm`` extension as CSV.
        This is pragmatic: Excel users often save as "CSV UTF-8" and the
        browser sends ``application/octet-stream`` — extension is more
        reliable than content-type.
        """
        name = (getattr(uploaded_file, 'name', '') or '').lower()
        if name.endswith(('.xlsx', '.xlsm')):
            return cls._parse_xlsx(uploaded_file, spec)
        return cls._parse_csv(uploaded_file, spec)

    # ── CSV ─────────────────────────────────────────────────────────────

    @classmethod
    def _parse_csv(cls, uploaded_file, spec: ImportSpec) -> ImportResult:
        raw = uploaded_file.read()
        if not isinstance(raw, (bytes, bytearray)):
            raw = str(raw).encode('utf-8')
        text = cls._decode(raw)
        if text is None:
            return ImportResult(
                rows=[], errors=[{
                    'row': 0, 'error':
                        'Could not decode the file. Try saving as '
                        'UTF-8 CSV and re-uploading.',
                }], columns=[], total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        # Auto-detect delimiter (comma vs tab vs semicolon).
        sample = text[:2048]
        delimiter = max([',', ';', '\t'], key=sample.count)
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows_iter = iter(reader)
        try:
            header = next(rows_iter)
        except StopIteration:
            return ImportResult(
                rows=[], errors=[{'row': 0, 'error': 'File is empty.'}],
                columns=[], total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        columns = cls._normalise_header(header)
        missing = cls._missing_required(columns, spec)
        if missing:
            return ImportResult(
                rows=[], errors=[{
                    'row': 1,
                    'error': (
                        f'Missing required column(s): {", ".join(missing)}. '
                        f'Got: {", ".join(columns)}.'
                    ),
                }], columns=columns, total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        return cls._build_result(rows_iter, columns, spec, start_row=2)

    # ── XLSX ────────────────────────────────────────────────────────────

    @classmethod
    def _parse_xlsx(cls, uploaded_file, spec: ImportSpec) -> ImportResult:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ImportResult(
                rows=[], errors=[{
                    'row': 0, 'error':
                        'openpyxl is not installed; cannot parse XLSX. '
                        'Upload as CSV instead.',
                }], columns=[], total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return ImportResult(
                rows=[], errors=[{'row': 0, 'error': 'Workbook is empty.'}],
                columns=[], total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        columns = cls._normalise_header(header_row)
        missing = cls._missing_required(columns, spec)
        if missing:
            return ImportResult(
                rows=[], errors=[{
                    'row': 1,
                    'error': (
                        f'Missing required column(s): {", ".join(missing)}. '
                        f'Got: {", ".join(columns)}.'
                    ),
                }], columns=columns, total_rows=0, accepted_rows=0, rejected_rows=0,
            )

        # openpyxl returns tuples — normalise to lists like CSV reader.
        def _iter_lists():
            for row in rows_iter:
                yield list(row) if row is not None else []

        return cls._build_result(_iter_lists(), columns, spec, start_row=2)

    # ── Shared body ─────────────────────────────────────────────────────

    @classmethod
    def _build_result(cls, rows_iter, columns, spec, *, start_row):
        rows: list[dict] = []
        errors: list[dict] = []
        total = 0
        accepted = 0
        rejected = 0

        for row_num, raw_row in enumerate(rows_iter, start=start_row):
            if _row_is_blank(raw_row):
                continue
            total += 1
            if total > spec.max_rows:
                errors.append({
                    'row': row_num,
                    'error': (
                        f'Upload exceeds the {spec.max_rows}-row cap. '
                        f'Split into smaller files.'
                    ),
                })
                rejected += 1
                break

            raw_dict = {
                col: (raw_row[i] if i < len(raw_row) else None)
                for i, col in enumerate(columns) if col
            }
            parsed, row_errs = cls._coerce_row(raw_dict, spec, row_num)
            if row_errs:
                errors.extend(row_errs)
                rejected += 1
            else:
                rows.append(parsed)
                accepted += 1

        return ImportResult(
            rows=rows, errors=errors, columns=columns,
            total_rows=total, accepted_rows=accepted, rejected_rows=rejected,
        )

    # ── Per-row coercion ────────────────────────────────────────────────

    @classmethod
    def _coerce_row(cls, raw: dict, spec: ImportSpec, row_num: int):
        """Coerce types per-column; collect per-cell errors."""
        errors: list[dict] = []
        out: dict[str, Any] = {}

        for col, val in raw.items():
            if col in spec.numeric_columns:
                parsed, err = cls._parse_decimal(val)
                if err:
                    errors.append({
                        'row': row_num, 'column': col,
                        'value': str(val), 'error': err,
                    })
                out[col] = parsed
            elif col in spec.date_columns:
                parsed, err = cls._parse_date(val)
                if err:
                    errors.append({
                        'row': row_num, 'column': col,
                        'value': str(val), 'error': err,
                    })
                out[col] = parsed
            else:
                out[col] = _clean_str(val)

        # Required-column presence check.
        for col in spec.required_columns:
            if out.get(col) in (None, '', Decimal('0')):
                # Decimal 0 is allowed for numeric amounts, but we flag
                # None/empty-string unless explicitly numeric column zero.
                if col in spec.numeric_columns and out.get(col) == Decimal('0'):
                    continue
                if out.get(col) in (None, ''):
                    errors.append({
                        'row': row_num, 'column': col,
                        'value': '', 'error': f'Required column {col!r} is empty.',
                    })

        return out, errors

    # ── Header / encoding utilities ─────────────────────────────────────

    @classmethod
    def _decode(cls, raw: bytes) -> str | None:
        for enc in cls.CSV_ENCODINGS:
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                continue
        return None

    @classmethod
    def _normalise_header(cls, header_row) -> list[str]:
        """Lowercase, strip whitespace, collapse inner spaces to ``_``."""
        out = []
        for cell in header_row:
            if cell is None:
                out.append('')
                continue
            s = str(cell).strip().lower()
            s = s.replace(' ', '_').replace('-', '_')
            # Strip characters other than word characters.
            s = ''.join(ch for ch in s if ch.isalnum() or ch == '_')
            out.append(s)
        return out

    @classmethod
    def _missing_required(cls, columns, spec: ImportSpec) -> list[str]:
        col_set = set(columns)
        return [c for c in spec.required_columns if c not in col_set]

    # ── Value parsers ───────────────────────────────────────────────────

    @classmethod
    def _parse_decimal(cls, value) -> tuple[Decimal, str | None]:
        if value is None or value == '':
            return Decimal('0'), None
        if isinstance(value, (int, float, Decimal)):
            try:
                return Decimal(str(value)), None
            except (InvalidOperation, ValueError):
                return Decimal('0'), f'Not a number: {value!r}'
        s = str(value).strip()
        if s.lower() in ('-', 'nil', 'n/a', 'none'):
            return Decimal('0'), None
        # Strip currency symbols and thousand separators.
        for tok in ('\u20a6', 'NGN', ',', ' '):
            s = s.replace(tok, '')
        negative = s.startswith('(') and s.endswith(')')
        if negative:
            s = s[1:-1]
        try:
            out = Decimal(s)
        except InvalidOperation:
            return Decimal('0'), f'Not a number: {value!r}'
        return (-out if negative else out), None

    @classmethod
    def _parse_date(cls, value) -> tuple[date | None, str | None]:
        if value is None or value == '':
            return None, None
        if isinstance(value, date):
            return value, None
        if isinstance(value, datetime):
            return value.date(), None
        s = str(value).strip()
        for fmt in cls.DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date(), None
            except ValueError:
                continue
        return None, f'Unrecognised date format: {value!r}'


# ── Internal helpers ───────────────────────────────────────────────────

def _row_is_blank(row) -> bool:
    return not any(
        (str(c).strip() if c is not None else '') for c in row
    )


def _clean_str(value) -> str:
    if value is None:
        return ''
    return str(value).strip()
