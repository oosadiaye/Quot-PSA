"""
Report rendering service — produces PDF, HTML, and Excel for IPSAS
financial statements and statutory schedules.

Backends
--------

  * **PDF**:  WeasyPrint if installed; otherwise a print-optimised
              HTML page is returned (browser handles PDF conversion).
  * **HTML**: Always available. Print-optimised CSS so "File → Print"
              in the browser produces a clean PDF matching the
              WeasyPrint output.
  * **Excel**: ``openpyxl``. Styled sheets with headers, totals,
              and auto-sized columns.

Design goals
------------
Every rendering target consumes the same ``ReportPayload`` protocol so
a new IPSAS report plugs in with zero boilerplate: define the dict,
call ``ReportRenderer.render()``, done.

Why not ``reportlab``? Lower-level — you write positioned text primitives,
which is excellent for invoices but overkill for multi-section statements
that already exist as structured dicts. HTML+CSS with WeasyPrint/browser
PDF is 10× faster to iterate on.
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

# ── Common HTML template (Jinja-free to keep dependencies minimal) ────

_PAGE_CSS = """
@page { size: A4; margin: 1.5cm 1.2cm 2cm 1.2cm; }
body {
    font-family: 'DejaVu Sans', 'Helvetica', 'Arial', sans-serif;
    font-size: 10pt;
    color: #111;
}
h1 { font-size: 16pt; margin: 0 0 2pt; }
h2 { font-size: 12pt; margin: 14pt 0 4pt; color: #0f4c75; border-bottom: 1pt solid #0f4c75; }
h3 { font-size: 10pt; margin: 8pt 0 3pt; }
.subtitle { font-size: 10pt; color: #555; margin-bottom: 10pt; }
table { width: 100%; border-collapse: collapse; margin: 4pt 0 8pt; }
th, td { padding: 3pt 6pt; text-align: left; }
th { border-bottom: 0.5pt solid #999; font-weight: 600; }
td.amount, th.amount { text-align: right; font-variant-numeric: tabular-nums; }
.total-row { font-weight: 700; border-top: 0.5pt solid #333; }
.grand-total-row { font-weight: 700; border-top: 1.5pt solid #0f4c75; background: #eaf2f9; }
.header-row { font-style: italic; color: #555; }
.cover-block {
    border: 1pt solid #ddd; padding: 10pt; margin: 8pt 0 16pt;
    background: #fafbfc;
}
.cover-block p { margin: 2pt 0; }
.meta { font-size: 9pt; color: #777; margin-top: 16pt; }
.signature-line {
    display: inline-block; width: 40%; border-bottom: 0.5pt solid #333;
    margin-top: 20pt; margin-right: 8%;
}
"""


class ReportRenderer:
    """Unified renderer for IPSAS statements + statutory exports."""

    # ── Public API ─────────────────────────────────────────────────────

    @classmethod
    def render(cls, report: dict, fmt: str) -> dict:
        """Render ``report`` in the requested format.

        Returns ``{'content': bytes|str, 'content_type': str,
        'suggested_filename': str}``. Callers wire this into an
        HTTP response.
        """
        fmt = (fmt or 'html').lower()
        if fmt == 'html':
            return cls._render_html(report)
        if fmt == 'pdf':
            return cls._render_pdf(report)
        if fmt in ('xlsx', 'excel'):
            return cls._render_xlsx(report)
        raise ValueError(
            f'Unsupported render format {fmt!r}. '
            f'Supported: html, pdf, xlsx.'
        )

    # ── HTML ───────────────────────────────────────────────────────────

    @classmethod
    def _render_html(cls, report: dict) -> dict:
        html = cls._build_html(report)
        stem = _safe_stem(report)
        return {
            'content':            html.encode('utf-8'),
            'content_type':       'text/html; charset=utf-8',
            'suggested_filename': f'{stem}.html',
        }

    @classmethod
    def _build_html(cls, report: dict) -> str:
        """Build a print-optimised HTML document for ``report``.

        Layout: cover-block with title/period/tenant + one section per
        top-level group of the report. Generic over the structure so
        SoFP, SoFPerformance, Cash Flow, etc. all work without per-
        statement templates.
        """
        title = report.get('title', 'Financial Report')
        subtitle = _build_subtitle(report)
        tenant = report.get('tenant_name') or report.get('tenant') or ''

        body_sections: list[str] = []

        # Render every dict-valued top-level key that has an 'items'
        # or 'total' inner key as a section. Plain scalar keys like
        # 'surplus_deficit' are picked up by _render_scalar_section.
        for key, value in report.items():
            if key in _SKIP_KEYS:
                continue
            if isinstance(value, dict) and _looks_like_section(value):
                body_sections.append(cls._render_section(key, value))
            elif isinstance(value, list) and key in _KNOWN_LIST_SECTIONS:
                body_sections.append(cls._render_rowlist_section(key, value))
            elif _is_scalar_amount(value):
                body_sections.append(cls._render_scalar_section(key, value))

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{_escape(title)}</title>
<style>{_PAGE_CSS}</style>
</head>
<body>
<h1>{_escape(title)}</h1>
<div class="subtitle">{_escape(subtitle)}</div>
<div class="cover-block">
  <p><strong>Entity:</strong> {_escape(tenant or 'Unspecified')}</p>
  <p><strong>Reporting period:</strong> {_escape(report.get('period_label', '') or str(report.get('period', '')))}</p>
  <p><strong>Currency:</strong> {_escape(report.get('currency', 'NGN'))}</p>
  <p><strong>Standard:</strong> {_escape(report.get('standard', '—'))}</p>
</div>
{''.join(body_sections)}
<div class="meta">
  <div class="signature-line"></div>
  <div class="signature-line"></div>
  <div>Prepared by &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Approved by</div>
  <p style="margin-top:10pt;">
    Generated on {_escape(str(report.get('generated_at', '')))}.
  </p>
</div>
</body>
</html>"""

    @classmethod
    def _render_section(cls, key: str, section: dict) -> str:
        label = _humanise_key(key)
        body = cls._render_section_body(section)
        return f'<h2>{_escape(label)}</h2>{body}'

    @classmethod
    def _render_section_body(cls, section: dict) -> str:
        """Render a nested IPSAS section (sub-groups with items + total)."""
        html: list[str] = []
        for sub_key, sub in section.items():
            if sub_key == 'total':
                continue
            if isinstance(sub, dict) and 'items' in sub:
                items = sub.get('items', []) or []
                sub_total = sub.get('total', Decimal('0'))
                html.append(
                    f'<h3>{_escape(_humanise_key(sub_key))}</h3>'
                    f'{_render_items_table(items)}'
                    f'{_render_total_row(sub_total, "Subtotal")}'
                )
            elif isinstance(sub, list):
                html.append(_render_items_table(sub))

        total = section.get('total')
        if total is not None:
            html.append(_render_total_row(total, 'Total', grand=True))
        return ''.join(html)

    @classmethod
    def _render_rowlist_section(cls, key: str, rows: list) -> str:
        label = _humanise_key(key)
        if not rows:
            return f'<h2>{_escape(label)}</h2><p><em>No records.</em></p>'
        # Derive columns from the first row's keys.
        columns = list(rows[0].keys())
        header = ''.join(
            f'<th class="{"amount" if _is_amount_column(c) else ""}">{_escape(str(c))}</th>'
            for c in columns
        )
        body = []
        for row in rows:
            cells = ''.join(
                f'<td class="{"amount" if _is_amount_column(c) else ""}">{_escape(_format_value(row.get(c, "")))}</td>'
                for c in columns
            )
            body.append(f'<tr>{cells}</tr>')
        return (
            f'<h2>{_escape(label)}</h2>'
            f'<table><thead><tr>{header}</tr></thead>'
            f'<tbody>{"".join(body)}</tbody></table>'
        )

    @classmethod
    def _render_scalar_section(cls, key: str, value: Any) -> str:
        label = _humanise_key(key)
        return (
            f'<h2>{_escape(label)}</h2>'
            f'<table><tr>'
            f'<td>{_escape(label)}</td>'
            f'<td class="amount"><strong>{_escape(_format_value(value))}</strong></td>'
            f'</tr></table>'
        )

    # ── PDF ───────────────────────────────────────────────────────────

    @classmethod
    def _render_pdf(cls, report: dict) -> dict:
        """Render to PDF via WeasyPrint when available; fall back to
        print-ready HTML when WeasyPrint or its system deps (Cairo,
        Pango) aren't present.

        The fallback HTML carries the same print CSS as the formal
        PDF, so "File → Print → Save as PDF" in Chrome produces a
        near-identical output. Many Nigerian AG offices operate this
        way already because their servers lack GTK libs.
        """
        html = cls._build_html(report)
        stem = _safe_stem(report)

        try:
            import weasyprint  # type: ignore
        except ImportError:
            return {
                'content':            html.encode('utf-8'),
                'content_type':       'text/html; charset=utf-8',
                'suggested_filename': f'{stem}.html',
                'fallback_reason':    (
                    'WeasyPrint not installed on this server — returning '
                    'print-ready HTML. Use your browser\'s "Print → Save '
                    'as PDF" to export.'
                ),
            }

        # WeasyPrint happy path.
        try:
            pdf_bytes = weasyprint.HTML(string=html).write_pdf()
        except Exception as exc:
            # GTK/Cairo runtime errors are common on stripped-down
            # Linux images. Fall back to HTML rather than 500.
            return {
                'content':            html.encode('utf-8'),
                'content_type':       'text/html; charset=utf-8',
                'suggested_filename': f'{stem}.html',
                'fallback_reason':    (
                    f'WeasyPrint PDF rendering failed ({exc.__class__.__name__}). '
                    'Returning print-ready HTML instead.'
                ),
            }

        return {
            'content':            pdf_bytes,
            'content_type':       'application/pdf',
            'suggested_filename': f'{stem}.pdf',
        }

    # ── Excel ──────────────────────────────────────────────────────────

    @classmethod
    def _render_xlsx(cls, report: dict) -> dict:
        """Render to .xlsx using openpyxl.

        Layout: one sheet per section. The cover sheet ("Summary")
        carries the period, entity, standard, and the top-level
        totals. A new sheet is added for each dict-valued section
        (SoFP → Assets/Liabilities/Net Assets sheets, etc.).
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        # ── Cover sheet ─────────────────────────────────────────────
        cover = wb.active
        cover.title = 'Summary'
        cover['A1'] = report.get('title', 'Financial Report')
        cover['A1'].font = Font(bold=True, size=14)
        cover['A2'] = _build_subtitle(report)
        cover['A2'].font = Font(italic=True, color='555555')

        row = 4
        for label, key in (
            ('Entity',            'tenant_name'),
            ('Reporting period',  'period_label'),
            ('Currency',          'currency'),
            ('Standard',          'standard'),
            ('Generated at',      'generated_at'),
        ):
            cover.cell(row=row, column=1, value=label).font = Font(bold=True)
            cover.cell(row=row, column=2, value=str(report.get(key, '')))
            row += 1

        # Top-level totals section on the cover.
        if 'totals' in report and isinstance(report['totals'], dict):
            row += 1
            cover.cell(row=row, column=1, value='Totals').font = Font(
                bold=True, color='0F4C75', size=12,
            )
            row += 1
            for k, v in report['totals'].items():
                cover.cell(row=row, column=1, value=_humanise_key(k))
                cover.cell(row=row, column=2, value=_to_number(v))
                cover.cell(row=row, column=2).number_format = '#,##0.00'
                row += 1

        cover.column_dimensions['A'].width = 32
        cover.column_dimensions['B'].width = 24

        # ── Per-section sheets ──────────────────────────────────────
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='0F4C75')
        total_font = Font(bold=True)
        total_fill = PatternFill('solid', fgColor='EAF2F9')

        for key, value in report.items():
            if key in _SKIP_KEYS or key == 'totals':
                continue
            if isinstance(value, dict) and _looks_like_section(value):
                cls._write_section_sheet(
                    wb, key, value,
                    header_font, header_fill, total_font, total_fill,
                )
            elif isinstance(value, list) and key in _KNOWN_LIST_SECTIONS:
                cls._write_rowlist_sheet(
                    wb, key, value, header_font, header_fill,
                )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stem = _safe_stem(report)
        return {
            'content':            buf.read(),
            'content_type':       (
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
            'suggested_filename': f'{stem}.xlsx',
        }

    @staticmethod
    def _write_section_sheet(
        wb, key: str, section: dict,
        header_font, header_fill, total_font, total_fill,
    ):
        # ``Font`` is imported locally so this staticmethod stays
        # self-contained (callers don't need to thread it through as
        # a parameter — would bloat the signature).
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        ws = wb.create_sheet(title=_sheet_name(key))

        row = 1
        ws.cell(row=row, column=1, value=_humanise_key(key)).font = Font(
            bold=True, size=14, color='0F4C75',
        )
        row += 2

        for sub_key, sub in section.items():
            if sub_key == 'total':
                continue
            if isinstance(sub, dict) and 'items' in sub:
                ws.cell(row=row, column=1, value=_humanise_key(sub_key)).font = Font(bold=True)
                row += 1
                # Header row
                cells = ['Code', 'Description', 'Amount (NGN)']
                for col, hv in enumerate(cells, start=1):
                    c = ws.cell(row=row, column=col, value=hv)
                    c.font = header_font
                    c.fill = header_fill
                row += 1
                for item in sub.get('items', []) or []:
                    ws.cell(row=row, column=1, value=str(item.get('code', '')))
                    ws.cell(row=row, column=2, value=str(item.get('name', '')))
                    ws.cell(row=row, column=3, value=_to_number(item.get('amount', 0)))
                    ws.cell(row=row, column=3).number_format = '#,##0.00'
                    row += 1
                subtotal = sub.get('total', 0)
                ws.cell(row=row, column=2, value='Subtotal').font = total_font
                c = ws.cell(row=row, column=3, value=_to_number(subtotal))
                c.font = total_font
                c.fill = total_fill
                c.number_format = '#,##0.00'
                row += 2

        total = section.get('total')
        if total is not None:
            ws.cell(row=row, column=2, value='Grand Total').font = Font(bold=True, size=12)
            c = ws.cell(row=row, column=3, value=_to_number(total))
            c.font = Font(bold=True, size=12)
            c.fill = total_fill
            c.number_format = '#,##0.00'

        # Column widths.
        for col, width in ((1, 14), (2, 40), (3, 20)):
            ws.column_dimensions[get_column_letter(col)].width = width

    @staticmethod
    def _write_rowlist_sheet(wb, key: str, rows: list, header_font, header_fill):
        from openpyxl.utils import get_column_letter
        ws = wb.create_sheet(title=_sheet_name(key))
        if not rows:
            ws['A1'] = f'(No rows for {_humanise_key(key)})'
            return
        columns = list(rows[0].keys())
        for col_idx, col in enumerate(columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=str(col))
            c.font = header_font
            c.fill = header_fill
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col in enumerate(columns, start=1):
                val = row.get(col, '')
                c = ws.cell(row=row_idx, column=col_idx, value=_to_number(val) if _is_amount_column(col) else str(val))
                if _is_amount_column(col):
                    c.number_format = '#,##0.00'

        # Auto-size columns — approximate via character count.
        for col_idx, col in enumerate(columns, start=1):
            max_len = max(
                [len(str(col))] + [len(str(r.get(col, ''))) for r in rows]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


# ── Module-level helpers ───────────────────────────────────────────────

# Keys in the top-level report dict that aren't themselves sections.
_SKIP_KEYS = frozenset({
    'title', 'subtitle', 'fiscal_year', 'period', 'period_label',
    'currency', 'standard', 'tenant_name', 'tenant', 'generated_at',
    'comparative',  # nested comparative payload has its own path
    'balance_check', 'reconciliation', 'totals',
})

# Keys whose list value carries heterogenous rows (exporter output).
_KNOWN_LIST_SECTIONS = frozenset({
    'items', 'rows', 'notes', 'lines',
})


def _looks_like_section(value: dict) -> bool:
    """Heuristic: a 'section' has either 'items' or sub-dicts with 'items'."""
    if 'items' in value or 'total' in value:
        return True
    return any(isinstance(v, dict) and 'items' in v for v in value.values())


def _is_scalar_amount(value) -> bool:
    return isinstance(value, (int, float, Decimal))


def _is_amount_column(column_name) -> bool:
    name = str(column_name).lower()
    return 'amount' in name or 'balance' in name or 'total' in name or 'ngn' in name


def _escape(value) -> str:
    """HTML-escape a string value."""
    text = str(value) if value is not None else ''
    return (
        text.replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
    )


def _format_value(value) -> str:
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return f'{value:,.2f}'
    if isinstance(value, float):
        return f'{value:,.2f}'
    return str(value)


def _humanise_key(key) -> str:
    return str(key).replace('_', ' ').title()


def _build_subtitle(report: dict) -> str:
    parts = []
    if report.get('standard'):
        parts.append(str(report['standard']))
    if report.get('period_label'):
        parts.append(f"Period {report['period_label']}")
    elif report.get('period'):
        parts.append(f"Period {report['period']}")
    return ' — '.join(parts)


def _safe_stem(report: dict) -> str:
    """Filename stem (no extension) for the rendered artefact."""
    base = str(
        report.get('title') or report.get('report_name') or 'report'
    ).lower().replace(' ', '-').replace('/', '-')
    # Strip any character that isn't safe in a filename.
    safe = ''.join(ch for ch in base if ch.isalnum() or ch in '-_')
    period = report.get('period_label') or report.get('period') or ''
    if period:
        safe = f'{safe}-{period}'
    return safe or 'report'


def _to_number(value) -> Any:
    """Coerce Decimals to float for Excel numeric cells; leave others
    as-is so openpyxl can stringify if needed."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return value
    try:
        return float(Decimal(str(value)))
    except Exception:
        return value


def _render_items_table(items: list) -> str:
    if not items:
        return '<p><em>No items.</em></p>'
    rows = []
    for it in items:
        cls_ = 'header-row' if it.get('is_header') else ''
        rows.append(
            f'<tr class="{cls_}">'
            f'<td>{_escape(it.get("code", ""))}</td>'
            f'<td>{_escape(it.get("name") or it.get("label") or "")}</td>'
            f'<td class="amount">{_escape(_format_value(it.get("amount", 0)))}</td>'
            f'</tr>'
        )
    return (
        '<table>'
        '<thead><tr><th>Code</th><th>Description</th><th class="amount">Amount</th></tr></thead>'
        f'<tbody>{"".join(rows)}</tbody></table>'
    )


def _render_total_row(total, label: str, grand: bool = False) -> str:
    cls_ = 'grand-total-row' if grand else 'total-row'
    return (
        f'<table><tr class="{cls_}">'
        f'<td></td><td>{_escape(label)}</td>'
        f'<td class="amount">{_escape(_format_value(total))}</td>'
        f'</tr></table>'
    )


def _sheet_name(key) -> str:
    """Excel sheet names are limited to 31 chars and can't contain ':*?/\\[]' ."""
    name = _humanise_key(key)
    for bad in (':', '*', '?', '/', '\\', '[', ']'):
        name = name.replace(bad, '')
    return name[:31]
