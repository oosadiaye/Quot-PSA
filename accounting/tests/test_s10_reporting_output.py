"""
Sprint-10 regression tests: report rendering + snapshot + IPSAS 19 models.

Fast-tier (no DB) coverage for:
  * ``ReportRenderer.render()`` — HTML, PDF (fallback path), Excel
  * ``ReportSnapshotService._serialise_and_hash()`` canonicalisation
  * ``_canonicalise()`` helper round-trips Decimals/dates
  * ``Provision.is_recognisable`` IPSAS 19 recognition gate
  * ``ContingentAsset.is_disclosable`` — probable-or-above only
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest


# =============================================================================
# ReportRenderer
# =============================================================================

class TestReportRendererHTML:

    def _sample_sofp(self):
        return {
            'title': 'Statement of Financial Position',
            'standard': 'IPSAS 1',
            'tenant_name': 'Delta State',
            'period_label': '2026-04',
            'currency': 'NGN',
            'assets': {
                'current': {
                    'items': [
                        {'code': '31', 'name': 'Cash', 'amount': Decimal('100000')},
                    ],
                    'total': Decimal('100000'),
                },
                'non_current': {
                    'items': [],
                    'total': Decimal('0'),
                },
                'total': Decimal('100000'),
            },
            'liabilities': {
                'current': {'items': [], 'total': Decimal('0')},
                'non_current': {'items': [], 'total': Decimal('0')},
                'total': Decimal('0'),
            },
            'net_assets': {
                'items': [{'code': '43', 'name': 'Accumulated Fund',
                           'amount': Decimal('100000')}],
                'total': Decimal('100000'),
            },
            'totals': {
                'assets': Decimal('100000'),
                'liabilities': Decimal('0'),
                'net_assets': Decimal('100000'),
            },
            'generated_at': '2026-04-17',
        }

    def test_html_contains_title_period_tenant(self):
        from accounting.services.report_rendering import ReportRenderer
        out = ReportRenderer.render(self._sample_sofp(), 'html')
        body = out['content'].decode('utf-8')
        assert 'Statement of Financial Position' in body
        assert '2026-04' in body
        assert 'Delta State' in body
        assert out['content_type'].startswith('text/html')
        assert out['suggested_filename'].endswith('.html')

    def test_html_renders_nested_sections_as_tables(self):
        from accounting.services.report_rendering import ReportRenderer
        out = ReportRenderer.render(self._sample_sofp(), 'html')
        body = out['content'].decode('utf-8')
        # The 'Assets' section should appear as an <h2>.
        assert '<h2>Assets</h2>' in body
        # Sub-group 'Current' as an <h3>.
        assert '<h3>Current</h3>' in body
        # Item row.
        assert 'Cash' in body
        # Amount rendered with thousands separator.
        assert '100,000.00' in body

    def test_html_escapes_dangerous_content(self):
        """XSS prevention — tenant names and item names go through escape."""
        from accounting.services.report_rendering import ReportRenderer
        payload = {
            'title': '<script>alert(1)</script>',
            'tenant_name': 'Delta & Sons',
            'items': [],
        }
        out = ReportRenderer.render(payload, 'html')
        body = out['content'].decode('utf-8')
        assert '<script>' not in body
        assert '&lt;script&gt;' in body
        assert 'Delta &amp; Sons' in body

    def test_unsupported_format_raises(self):
        from accounting.services.report_rendering import ReportRenderer
        with pytest.raises(ValueError, match='Unsupported render format'):
            ReportRenderer.render({}, 'docx')


class TestReportRendererPDFFallback:
    """WeasyPrint isn't installed; the PDF path should fall back to HTML
    with a ``fallback_reason`` field that callers can surface."""

    def test_pdf_falls_back_to_html_when_weasyprint_missing(self):
        from accounting.services.report_rendering import ReportRenderer
        out = ReportRenderer.render(
            {'title': 'Test', 'tenant_name': 'X', 'totals': {}},
            'pdf',
        )
        # Either true PDF (if somehow installed) or HTML fallback.
        if out['content_type'] == 'application/pdf':
            assert out['suggested_filename'].endswith('.pdf')
        else:
            assert out['content_type'].startswith('text/html')
            assert out['suggested_filename'].endswith('.html')
            assert 'fallback_reason' in out


class TestReportRendererExcel:

    def test_excel_produces_xlsx_bytes(self):
        from accounting.services.report_rendering import ReportRenderer
        out = ReportRenderer.render({
            'title': 'WHT Schedule', 'tenant_name': 'Delta State',
            'period_label': '2026-04',
            'totals': {'total_wht_amount': Decimal('100000')},
            'rows': [
                {'SN': 1, 'Beneficiary Name': 'Acme',
                 'WHT Amount (NGN)': Decimal('5000')},
            ],
        }, 'xlsx')
        # .xlsx files start with the PK magic number (they're ZIPs).
        assert out['content'][:2] == b'PK'
        assert out['content_type'].endswith('spreadsheetml.sheet')
        assert out['suggested_filename'].endswith('.xlsx')

    def test_excel_has_summary_sheet_and_section_sheets(self):
        from accounting.services.report_rendering import ReportRenderer
        import io
        from openpyxl import load_workbook

        payload = {
            'title': 'Cash Flow Statement',
            'tenant_name': 'Delta State',
            'standard': 'IPSAS 2',
            'period_label': '2026-04',
            'totals': {
                'net_operating': Decimal('1000000'),
                'net_investing': Decimal('-500000'),
                'net_financing': Decimal('0'),
            },
            'operating_activities': {
                'items': [
                    {'code': 'OP-1', 'name': 'Tax receipts', 'amount': Decimal('1500000')},
                ],
                'total': Decimal('1500000'),
            },
        }
        out = ReportRenderer.render(payload, 'xlsx')
        wb = load_workbook(io.BytesIO(out['content']))
        assert 'Summary' in wb.sheetnames
        # Section sheet created for the operating_activities key.
        # Sheet name is humanised, 31-char trimmed.
        sheet_names = ' '.join(wb.sheetnames).lower()
        assert 'operating' in sheet_names

    def test_excel_formats_decimal_columns(self):
        """Amount cells should carry the '#,##0.00' number format."""
        from accounting.services.report_rendering import ReportRenderer
        import io
        from openpyxl import load_workbook

        payload = {
            'title': 'WHT', 'tenant_name': 'X',
            'totals': {'total_wht': Decimal('5000')},
            'rows': [
                {'SN': 1, 'Beneficiary Name': 'Alpha',
                 'WHT Amount (NGN)': Decimal('5000')},
            ],
        }
        out = ReportRenderer.render(payload, 'xlsx')
        wb = load_workbook(io.BytesIO(out['content']))
        # Find the 'Rows' sheet.
        sheet = wb['Rows']
        # Amount column is the third header.
        assert sheet.cell(row=1, column=3).value == 'WHT Amount (NGN)'
        assert sheet.cell(row=2, column=3).number_format == '#,##0.00'


# =============================================================================
# ReportSnapshotService
# =============================================================================

class TestSnapshotSerialisation:

    def test_canonicalise_decimal_to_str(self):
        from accounting.services.report_snapshot import _canonicalise
        assert _canonicalise(Decimal('100.50')) == '100.50'

    def test_canonicalise_date_to_iso(self):
        from accounting.services.report_snapshot import _canonicalise
        assert _canonicalise(date(2026, 4, 17)) == '2026-04-17'

    def test_canonicalise_nested(self):
        from accounting.services.report_snapshot import _canonicalise
        payload = {
            'period': date(2026, 4, 1),
            'items': [
                {'amount': Decimal('100'), 'name': 'A'},
                {'amount': Decimal('200'), 'name': 'B'},
            ],
        }
        result = _canonicalise(payload)
        assert result == {
            'period': '2026-04-01',
            'items': [
                {'amount': '100', 'name': 'A'},
                {'amount': '200', 'name': 'B'},
            ],
        }

    def test_hash_is_deterministic_over_key_order(self):
        """Two payloads with the same content but different key order
        produce the same hash because we sort keys at serialise time."""
        from accounting.services.report_snapshot import ReportSnapshotService
        p1 = {'a': 1, 'b': 2, 'c': 3}
        p2 = {'c': 3, 'a': 1, 'b': 2}
        _, h1 = ReportSnapshotService._serialise_and_hash(p1)
        _, h2 = ReportSnapshotService._serialise_and_hash(p2)
        assert h1 == h2

    def test_hash_changes_when_content_changes(self):
        from accounting.services.report_snapshot import ReportSnapshotService
        _, h1 = ReportSnapshotService._serialise_and_hash({'a': 1})
        _, h2 = ReportSnapshotService._serialise_and_hash({'a': 2})
        assert h1 != h2

    def test_hash_is_64_char_hex(self):
        """SHA-256 hex digest is always 64 characters."""
        from accounting.services.report_snapshot import ReportSnapshotService
        _, h = ReportSnapshotService._serialise_and_hash({'x': 'y'})
        assert len(h) == 64
        assert all(c in '0123456789abcdef' for c in h)


# =============================================================================
# Provision (IPSAS 19)
# =============================================================================

class TestProvisionRecognitionGate:

    def _provision(self, *, amount=Decimal('100000'), likelihood='PROBABLE'):
        """Build an unsaved Provision instance for pure-logic testing."""
        from accounting.models import Provision
        return Provision(
            reference='PROV-TEST-1',
            category='LITIGATION',
            title='Test Provision',
            description='Test',
            amount=amount,
            likelihood=likelihood,
            recognition_date=date(2026, 4, 1),
        )

    def test_probable_with_positive_amount_is_recognisable(self):
        p = self._provision()
        assert p.is_recognisable is True

    def test_certain_likelihood_is_recognisable(self):
        p = self._provision(likelihood='CERTAIN')
        assert p.is_recognisable is True

    def test_possible_likelihood_not_recognisable(self):
        """IPSAS 19 ¶22: recognition requires "probable" (>50%). A merely
        possible obligation is disclosed as a contingent liability, not
        recognised as a provision."""
        p = self._provision(likelihood='POSSIBLE')
        assert p.is_recognisable is False

    def test_remote_likelihood_not_recognisable(self):
        p = self._provision(likelihood='REMOTE')
        assert p.is_recognisable is False

    def test_zero_amount_not_recognisable(self):
        """Reliable measurement requires a non-zero estimate."""
        p = self._provision(amount=Decimal('0'))
        assert p.is_recognisable is False

    def test_negative_amount_not_recognisable(self):
        """Defensive: a negative estimate is never recognisable."""
        p = self._provision(amount=Decimal('-100'))
        assert p.is_recognisable is False


class TestContingentAssetDisclosure:

    def _asset(self, *, likelihood):
        from accounting.models import ContingentAsset
        return ContingentAsset(
            reference='CA-1',
            title='Test',
            description='Test',
            likelihood=likelihood,
            arising_date=date(2026, 4, 1),
        )

    def test_probable_is_disclosable(self):
        assert self._asset(likelihood='PROBABLE').is_disclosable is True

    def test_certain_is_disclosable(self):
        assert self._asset(likelihood='CERTAIN').is_disclosable is True

    def test_possible_not_disclosable(self):
        """IPSAS 19 ¶39: contingent assets disclosed ONLY when probable."""
        assert self._asset(likelihood='POSSIBLE').is_disclosable is False

    def test_remote_not_disclosable(self):
        assert self._asset(likelihood='REMOTE').is_disclosable is False
